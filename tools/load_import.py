#!/usr/bin/env python3
"""Load the reviewed import sheet into the RSM B2B backend.

  RSM_ADMIN_PASS=... python3 tools/load_import.py <import.xlsx> [--dry-run]

One adminSaveProduct call per row. The call is idempotent: re-running updates
the same SKU rather than creating a duplicate, so a partial run can simply be
repeated. Apparel outside Headwear gets the standard size range unless the
sheet names sizes explicitly.

Images are referenced as assets/products/<SKU>.webp, the same relative path the
existing catalogue uses; tools/import_images.py puts the files there.
"""
import json
import os
import sys
import urllib.request

import openpyxl

API = ('https://script.google.com/macros/s/AKfycbyDezChvk8YkvxbdaPMB0W5sKK1z'
       'nGFH7F6B9T0ficleUdVTPGk22tPD-MI_hZeelaf/exec')
TOKEN = 'rsm_XkCA0HS327rSxemHHRHIymHolJcf'
DEFAULT_SIZES = ['XS', 'S', 'M', 'L', 'XL', '2XL', '3XL']


def call(fn, admin_pass, **kw):
    body = json.dumps({'fn': fn, 'token': TOKEN, 'admin_pass': admin_pass, **kw}).encode()
    req = urllib.request.Request(API, body, {'Content-Type': 'text/plain'})
    return json.loads(urllib.request.urlopen(req, timeout=120).read())


def main():
    path = sys.argv[1]
    dry = '--dry-run' in sys.argv
    admin_pass = os.environ.get('RSM_ADMIN_PASS')
    if not admin_pass and not dry:
        raise SystemExit('Set RSM_ADMIN_PASS first.')

    ws = openpyxl.load_workbook(path)['import']
    head = [c.value for c in ws[1]]
    col = {name: i for i, name in enumerate(head)}

    ok, failed = 0, []
    for r in ws.iter_rows(min_row=2, values_only=True):
        sku = r[col['sku']]
        if not sku:
            continue

        tiers = []
        for n in range(1, 7):
            q, p = r[col[f'tier{n}_qty']], r[col[f'tier{n}_price']]
            if q and p:
                tiers.append({'min_qty': int(q),
                              'max_qty': '', 'unit_price': float(p)})
        for i in range(len(tiers) - 1):
            tiers[i]['max_qty'] = tiers[i + 1]['min_qty'] - 1

        sizes = str(r[col['sizes']] or '').split()
        if not sizes and r[col['category']] == 'Apparel' \
                and r[col['subcategory']] != 'Headwear':
            sizes = DEFAULT_SIZES

        product = {
            'sku': sku,
            'name': r[col['name']],
            'category': r[col['category']],
            'subcategory': r[col['subcategory']],
            'description': r[col['description']] or '',
            'moq': int(r[col['moq']]),
            'gst_rate': float(r[col['gst_rate']] or 0),
            'lead_time_days': int(r[col['lead_time_days']] or 21),
            'image': f'assets/products/{sku}.webp',
            'sizes': sizes,
            'tiers': tiers,
            'related_skus': [],
            'active': True,
        }

        if dry:
            print(f"{sku}  {product['name'][:38]:38s} MOQ {product['moq']:4d} "
                  f"{len(tiers)} tiers  {len(sizes)} sizes")
            ok += 1
            continue

        try:
            res = call('adminSaveProduct', admin_pass, product=product)
            if not res.get('ok'):
                raise RuntimeError(res.get('error'))
            print(f"{sku}  {'created' if res.get('created') else 'updated'}")
            ok += 1
        except Exception as err:
            print(f'{sku}  FAILED  {err}')
            failed.append(sku)

    print(f'\n{ok} loaded, {len(failed)} failed')
    if failed:
        print('failed:', ', '.join(failed))
        sys.exit(1)


if __name__ == '__main__':
    main()
