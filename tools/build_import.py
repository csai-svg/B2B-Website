#!/usr/bin/env python3
"""Turn Sujith's workbook into a reviewable import sheet for the RSM B2B store.

  python3 tools/build_import.py <workbook.xlsx> <out.xlsx>

Reads the resolved values on 'Sheet12' (the 'Product file' tab holds formulas),
attaches the SKU from tools/row_to_sku.json, and writes one row per product with
the price tiers the store needs.

Rules applied here, all visible in the output so they can be argued with:
  * price breaks are MOQ, 50, 100, 200, 500, 1000; any break below the product's
    own MOQ is dropped, and the first surviving break is moved to the MOQ so the
    store's "first tier starts at the MOQ" rule holds
  * prices are rounded to the nearest rupee
  * a break priced 'NA' or blank is dropped
  * tax 0.05 in the sheet means 5% GST
  * category and subcategory are proposed from the product name and reuse the
    subcategories already in the store
"""
import json
import os
import re
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill

HERE = os.path.dirname(os.path.abspath(__file__))
# (price column, quantity the band starts at). None means "starts at the MOQ".
# Each price column is followed by its own margin column, which is used below to
# check the price actually belongs to that band.
BREAKS = [(11, None), (13, 50), (15, 100), (17, 200), (19, 500), (21, 1000)]

RULES = [
    (r'\bt-?shirt\b', 'Apparel', 'T-Shirts'),
    (r'\bpolo\b|\bjersey\b(?!.*cap)|\bshirt\b', 'Apparel', 'Shirts'),
    (r'hoodie|sweatshirt|funnel neck', 'Apparel', 'Hoodies'),
    (r'jacket', 'Apparel', 'Jackets'),
    (r'\bcap\b|bucket hat|visor', 'Apparel', 'Headwear'),
    (r'bottle|sipper', 'Drinkware', 'Bottles'),
    (r'\bmug\b|tumbler|flask', 'Drinkware', 'Mugs & tumblers'),
    (r'backpack|back pack', 'Travel', 'Bag pack'),
    (r'duffle', 'Travel', 'Duffle bag'),
    (r'tote', 'Travel', 'Tote bags'),
    (r'sling|belt bag|fanny pack|crossbody|cross body|pouch', 'Travel', 'Laptop handbag'),
]

HEAD = ['sku', 'name', 'brand', 'category', 'subcategory', 'gender', 'moq',
        'gst_rate', 'lead_time_days', 'sizes', 'image_file', 'description',
        'tier1_qty', 'tier1_price', 'tier2_qty', 'tier2_price',
        'tier3_qty', 'tier3_price', 'tier4_qty', 'tier4_price',
        'tier5_qty', 'tier5_price', 'tier6_qty', 'tier6_price',
        'mrp', 'cost', 'source_row', 'NEEDS REVIEW']


def classify(name):
    low = name.lower()
    for pattern, cat, sub in RULES:
        if re.search(pattern, low):
            return cat, sub
    return 'Utilities', 'Accessories'


def money(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        if not v or v.upper() == 'NA':
            return None
        try:
            v = float(v.replace(',', ''))
        except ValueError:
            return None
    return int(round(float(v)))


def main():
    book, out = sys.argv[1], sys.argv[2]
    skus = {int(k): v for k, v in
            json.load(open(os.path.join(HERE, 'row_to_sku.json'))).items()}

    ws = openpyxl.load_workbook(book)['Sheet12']
    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = 'import'
    sh.append(HEAD)
    for c in sh[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill('solid', fgColor='F4F8FB')
    sh.freeze_panes = 'A2'

    warn = PatternFill('solid', fgColor='FDE7E5')
    count = 0

    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        row_no = i + 2
        if row_no not in skus or not r[2]:
            continue

        name = ' '.join(str(r[2]).split())
        moq = int(float(r[10] or 0))
        cat, sub = classify(name)

        cost = money(r[9])
        tiers, notes = [], []
        for col, start in BREAKS:
            price = money(r[col])
            if price is None:
                continue
            qty = moq if start is None or start < moq else start

            # The sheet carries a margin beside every price. (price - cost) /
            # price should reproduce it; where it does not, the price cell has
            # been overridden by hand and the row needs a human eye.
            margin = r[col + 1]
            if cost and isinstance(margin, (int, float)):
                implied = (price - cost) / price
                if abs(implied - float(margin)) > 0.01:
                    notes.append(f'price at {qty}+ disagrees with its margin '
                                 f'(sheet says {float(margin):.3f}, '
                                 f'₹{price} implies {implied:.3f})')

            tiers = [t for t in tiers if t[0] != qty]
            tiers.append((qty, price))
        tiers.sort()
        if tiers and tiers[0][0] != moq:
            tiers[0] = (moq, tiers[0][1])
        if not tiers:
            notes.append('no usable price')
        if len(tiers) < 2:
            notes.append('only one price break')
        if any(tiers[i][1] < tiers[i + 1][1] for i in range(len(tiers) - 1)):
            notes.append('price rises with quantity')
        if cat == 'Apparel' and sub != 'Headwear':
            notes.append('sizes needed')

        flat = []
        for q, p in tiers[:6]:
            flat += [q, p]
        flat += [None] * (12 - len(flat))

        sh.append([skus[row_no], name, str(r[3] or '').strip(), cat, sub,
                   str(r[5] or '').strip(), moq, round(float(r[7] or 0) * 100, 2),
                   int(re.sub(r'\D', '', str(r[23] or '21')) or 21),
                   '', skus[row_no] + '.png', str(r[4] or '').strip()]
                  + flat + [money(r[6]), money(r[9]), row_no, '; '.join(notes)])
        if notes:
            for c in sh[sh.max_row]:
                c.fill = warn
        count += 1

    widths = {'A': 16, 'B': 40, 'C': 12, 'D': 11, 'E': 17, 'F': 9, 'L': 60, 'AB': 26}
    for col, w in widths.items():
        sh.column_dimensions[col].width = w

    wb.save(out)
    print(f'{count} products -> {out}')


if __name__ == '__main__':
    main()
